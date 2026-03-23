"""
Generates a monthly Habit Tracker / To-Do List Excel spreadsheet.
- Rows = habits/tasks
- Columns = days of the month (1–31)
- Type an X (or any character) in a day cell to mark it done
- Conditional formatting turns completed cells green
- Includes a progress % column and summary stats
"""

import calendar
import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Configuration ────────────────────────────────────────────────────────────
YEAR  = 2026
MONTH = 4          # Change to desired month (1–12)
MONTH_NAME = calendar.month_name[MONTH]
DAYS_IN_MONTH = calendar.monthrange(YEAR, MONTH)[1]

# Sample habits / tasks – feel free to edit
HABITS = [
    ("Morning Routine", "🌅 Morning"),
    ("Exercise / Workout", "💪 Health"),
    ("Drink 8 Glasses of Water", "💧 Health"),
    ("Read for 30 Minutes", "📚 Learning"),
    ("Meditate / Mindfulness", "🧘 Wellness"),
    ("Healthy Eating", "🥗 Health"),
    ("No Social Media Before 9am", "📵 Focus"),
    ("Journal / Gratitude", "✍️ Wellness"),
    ("Work / Study Goals", "🎯 Productivity"),
    ("Evening Walk / Stretch", "🚶 Health"),
    ("Sleep by 11pm", "😴 Wellness"),
    ("-- TO-DO SECTION --", ""),
    ("Task 1", "📋 To-Do"),
    ("Task 2", "📋 To-Do"),
    ("Task 3", "📋 To-Do"),
    ("Task 4", "📋 To-Do"),
    ("Task 5", "📋 To-Do"),
]

# ── Colours ──────────────────────────────────────────────────────────────────
HEADER_BG      = "1F3864"   # dark navy
HEADER_FG      = "FFFFFF"
SUBHEADER_BG   = "2E75B6"   # blue
SUBHEADER_FG   = "FFFFFF"
WEEKEND_COL    = "D9E1F2"   # light blue tint for weekend columns
DONE_FILL      = "70AD47"   # green when marked
DONE_FONT      = "FFFFFF"
ALT_ROW        = "F2F2F2"   # light grey alternating rows
SEPARATOR_BG   = "BDD7EE"   # section divider
PROGRESS_BG    = "FFF2CC"   # yellow tint for progress col
CATEGORY_COLORS = {
    "🌅 Morning":       "FFE699",
    "💪 Health":        "C6EFCE",
    "💧 Health":        "C6EFCE",
    "🥗 Health":        "C6EFCE",
    "🚶 Health":        "C6EFCE",
    "📚 Learning":      "DDEBF7",
    "🧘 Wellness":      "E2EFDA",
    "😴 Wellness":      "E2EFDA",
    "✍️ Wellness":      "E2EFDA",
    "📵 Focus":         "FCE4D6",
    "🎯 Productivity":  "EDEDED",
    "📋 To-Do":         "FFF2CC",
}

# ── Helper styles ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def bold_font(color="000000", size=11):
    return Font(bold=True, color=color, size=size)

def normal_font(color="000000", size=10):
    return Font(color=color, size=size)

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Monthly Tracker ──────────────────────────────────────────────────
ws = wb.active
ws.title = f"{MONTH_NAME} {YEAR}"

FIRST_DATA_ROW = 4          # row where habit rows start
HABIT_COL      = 2          # column B = habit name
CATEGORY_COL   = 3          # column C = category
FIRST_DAY_COL  = 4          # column D = day 1
LAST_DAY_COL   = FIRST_DAY_COL + DAYS_IN_MONTH - 1
PROGRESS_COL   = LAST_DAY_COL + 1
NOTES_COL      = PROGRESS_COL + 1

# Row 1: Big title
ws.merge_cells(start_row=1, start_column=1,
               end_row=1, end_column=NOTES_COL)
title_cell = ws.cell(row=1, column=1,
                     value=f"📅  {MONTH_NAME.upper()} {YEAR}  —  HABIT TRACKER & TO-DO LIST")
title_cell.font      = Font(bold=True, color=HEADER_FG, size=16)
title_cell.fill      = fill(HEADER_BG)
title_cell.alignment = center()
ws.row_dimensions[1].height = 36

# Row 2: instruction
ws.merge_cells(start_row=2, start_column=1,
               end_row=2, end_column=NOTES_COL)
instr = ws.cell(row=2, column=1,
                value="✅  Type  X  in any day cell to mark it complete  —  cell turns green automatically")
instr.font      = Font(italic=True, color="595959", size=10)
instr.fill      = fill("EBF3FB")
instr.alignment = center()
ws.row_dimensions[2].height = 20

# Row 3: column headers
ws.cell(row=3, column=HABIT_COL, value="Habit / Task").font = bold_font(HEADER_FG)
ws.cell(row=3, column=HABIT_COL).fill = fill(SUBHEADER_BG)
ws.cell(row=3, column=HABIT_COL).alignment = center(wrap=True)

ws.cell(row=3, column=CATEGORY_COL, value="Category").font = bold_font(HEADER_FG)
ws.cell(row=3, column=CATEGORY_COL).fill = fill(SUBHEADER_BG)
ws.cell(row=3, column=CATEGORY_COL).alignment = center()

for d in range(1, DAYS_IN_MONTH + 1):
    col = FIRST_DAY_COL + d - 1
    weekday = datetime.date(YEAR, MONTH, d).strftime("%a")  # Mon, Tue…
    cell = ws.cell(row=3, column=col, value=f"{d}\n{weekday}")
    cell.font      = bold_font(HEADER_FG, size=9)
    cell.alignment = center(wrap=True)
    is_weekend = datetime.date(YEAR, MONTH, d).weekday() >= 5
    cell.fill = fill("3A6BBF" if not is_weekend else "5B9BD5")

ws.cell(row=3, column=PROGRESS_COL, value="Progress\n(%)").font = bold_font(HEADER_FG)
ws.cell(row=3, column=PROGRESS_COL).fill = fill(SUBHEADER_BG)
ws.cell(row=3, column=PROGRESS_COL).alignment = center(wrap=True)

ws.cell(row=3, column=NOTES_COL, value="Notes").font = bold_font(HEADER_FG)
ws.cell(row=3, column=NOTES_COL).fill = fill(SUBHEADER_BG)
ws.cell(row=3, column=NOTES_COL).alignment = center()

ws.row_dimensions[3].height = 32

# ── Habit rows ────────────────────────────────────────────────────────────────
for idx, (habit, category) in enumerate(HABITS):
    row = FIRST_DATA_ROW + idx
    is_separator = habit.startswith("--")
    alt = (idx % 2 == 1)

    if is_separator:
        # Section divider row
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=NOTES_COL)
        sep = ws.cell(row=row, column=1, value="  📋  TO-DO LIST")
        sep.font      = bold_font(HEADER_FG, size=11)
        sep.fill      = fill(SUBHEADER_BG)
        sep.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24
        continue

    ws.row_dimensions[row].height = 22

    # Habit name
    hcell = ws.cell(row=row, column=HABIT_COL, value=habit)
    hcell.font      = Font(size=10, bold=False)
    hcell.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1)
    row_bg = CATEGORY_COLORS.get(category, ALT_ROW if alt else "FFFFFF")
    hcell.fill   = fill(row_bg)
    hcell.border = thin_border()

    # Category
    ccell = ws.cell(row=row, column=CATEGORY_COL, value=category)
    ccell.font      = Font(size=9, italic=True, color="595959")
    ccell.alignment = center()
    ccell.fill      = fill(row_bg)
    ccell.border    = thin_border()

    # Day cells (D … D+30)
    day_range_start = get_column_letter(FIRST_DAY_COL) + str(row)
    day_range_end   = get_column_letter(LAST_DAY_COL)  + str(row)

    for d in range(1, DAYS_IN_MONTH + 1):
        col = FIRST_DAY_COL + d - 1
        dcell = ws.cell(row=row, column=col, value="")
        dcell.alignment = center()
        is_weekend = datetime.date(YEAR, MONTH, d).weekday() >= 5
        dcell.fill   = fill(WEEKEND_COL if is_weekend else row_bg)
        dcell.border = thin_border()

    # Progress formula: =COUNTIF(D{row}:AH{row},"<>")/DAYS_IN_MONTH
    first_day_letter = get_column_letter(FIRST_DAY_COL)
    last_day_letter  = get_column_letter(LAST_DAY_COL)
    pcell = ws.cell(
        row=row, column=PROGRESS_COL,
        value=f'=COUNTIF({first_day_letter}{row}:{last_day_letter}{row},"<>")/{DAYS_IN_MONTH}'
    )
    pcell.number_format = "0%"
    pcell.alignment     = center()
    pcell.fill          = fill(PROGRESS_BG)
    pcell.font          = Font(size=10, bold=True, color="7F6000")
    pcell.border        = thin_border()

    # Notes
    ncell = ws.cell(row=row, column=NOTES_COL, value="")
    ncell.fill   = fill(row_bg)
    ncell.border = thin_border()

# ── Conditional formatting: X → green ────────────────────────────────────────
green_fill = PatternFill("solid", fgColor=DONE_FILL)
green_font = Font(bold=True, color=DONE_FONT, size=10)

cf_range = (
    f"{get_column_letter(FIRST_DAY_COL)}{FIRST_DATA_ROW}:"
    f"{get_column_letter(LAST_DAY_COL)}{FIRST_DATA_ROW + len(HABITS) - 1}"
)

ws.conditional_formatting.add(
    cf_range,
    CellIsRule(operator="notEqual", formula=['"  "'],
               fill=green_fill, font=green_font)
)
# Also catch actual non-blank cells
ws.conditional_formatting.add(
    cf_range,
    FormulaRule(
        formula=[f'{get_column_letter(FIRST_DAY_COL)}{FIRST_DATA_ROW}<>""'],
        fill=green_fill,
        font=green_font,
    )
)

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(1)].width             = 1.5   # spacer
ws.column_dimensions[get_column_letter(HABIT_COL)].width     = 28
ws.column_dimensions[get_column_letter(CATEGORY_COL)].width  = 16
for d in range(DAYS_IN_MONTH):
    ws.column_dimensions[get_column_letter(FIRST_DAY_COL + d)].width = 4.5
ws.column_dimensions[get_column_letter(PROGRESS_COL)].width  = 10
ws.column_dimensions[get_column_letter(NOTES_COL)].width     = 22

# Freeze panes so habit names stay visible when scrolling
ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=FIRST_DAY_COL)

# ── Sheet 2: Instructions ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Instructions")
instructions = [
    ("HOW TO USE THIS HABIT TRACKER", True, 14, HEADER_BG, HEADER_FG),
    ("", False, 11, "FFFFFF", "000000"),
    ("1.  MARK A HABIT DONE", True, 12, SUBHEADER_BG, SUBHEADER_FG),
    ("    Click any day cell next to a habit and type  X  (or any character).", False, 11, "FFFFFF", "000000"),
    ("    The cell will automatically turn GREEN.", False, 11, "FFFFFF", "000000"),
    ("    To unmark, simply delete the cell contents.", False, 11, "FFFFFF", "000000"),
    ("", False, 11, "FFFFFF", "000000"),
    ("2.  PROGRESS COLUMN", True, 12, SUBHEADER_BG, SUBHEADER_FG),
    ("    The '% Progress' column calculates how many days you completed the habit", False, 11, "FFFFFF", "000000"),
    ("    out of the total days in the month. It updates automatically.", False, 11, "FFFFFF", "000000"),
    ("", False, 11, "FFFFFF", "000000"),
    ("3.  ADDING YOUR OWN HABITS / TASKS", True, 12, SUBHEADER_BG, SUBHEADER_FG),
    ("    Replace the sample habit names in column B with your own.", False, 11, "FFFFFF", "000000"),
    ("    Update the Category column (C) to match.", False, 11, "FFFFFF", "000000"),
    ("    The green conditional formatting applies to the whole day grid automatically.", False, 11, "FFFFFF", "000000"),
    ("", False, 11, "FFFFFF", "000000"),
    ("4.  NOTES", True, 12, SUBHEADER_BG, SUBHEADER_FG),
    ("    Use the Notes column (last column) to add any reminders or comments.", False, 11, "FFFFFF", "000000"),
    ("", False, 11, "FFFFFF", "000000"),
    ("5.  WEEKEND COLUMNS", True, 12, SUBHEADER_BG, SUBHEADER_FG),
    ("    Saturday and Sunday columns are tinted blue so they're easy to spot.", False, 11, "FFFFFF", "000000"),
    ("", False, 11, "FFFFFF", "000000"),
    ("TIP: Use  Ctrl + Home  to jump back to the top of the tracker sheet.", True, 11, "FFF2CC", "7F6000"),
]

for r, (text, bold, size, bg, fg) in enumerate(instructions, start=1):
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell = ws2.cell(row=r, column=1, value=text)
    cell.font      = Font(bold=bold, size=size, color=fg)
    cell.fill      = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[r].height = 20 if text else 8

for col in range(1, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 18

# ── Save ──────────────────────────────────────────────────────────────────────
filename = f"Habit_Tracker_{MONTH_NAME}_{YEAR}.xlsx"
wb.save(filename)
print(f"✅  Saved: {filename}")
